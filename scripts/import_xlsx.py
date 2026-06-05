#!/usr/bin/env python3
"""
One-time importer: load Financial Tracker 2026.xlsx into Supabase Postgres.

Setup once:
    python -m pip install "psycopg[binary]" openpyxl
    # put DATABASE_URL in .env.local (Supabase -> Settings -> Database -> Connection string)

Run:
    python scripts/import_xlsx.py              # ensures schema+seed, then imports
    python scripts/import_xlsx.py --dry-run    # parse + print counts, write nothing

Re-running is safe: reference tables (wallets/categories) are kept; the historical
data tables are truncated and re-imported.
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "source-data" / "Financial Tracker 2026.xlsx"
YEAR = 2026

INVESTMENT_BUCKETS = {"Forex Yen", "Stock", "Bonds"}
SAVING_BUCKETS = {"Dana Darurat", "Dana Pensiun", "Jepang", "Rumah"}
MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]


# ----------------------------- helpers -----------------------------

def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if path.exists():
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def to_date(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()[:10]
        try:
            return datetime.fromisoformat(s).date().isoformat()
        except ValueError:
            return None
    return None


def to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(round(abs(float(v))))
    except (TypeError, ValueError):
        return None


def truthy(v):
    if isinstance(v, bool):
        return v
    if v is None or v == "":
        return None
    if isinstance(v, str):
        s = v.strip().lower()
        if s == "true":
            return True
        if s == "false":
            return False
    return None


def text(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ----------------------------- parsing -----------------------------

def parse_workbook(xlsx: Path):
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    transactions = []   # dicts: occurred_on, type, amount, description, category(kind,name), src_wallet, dst_wallet
    balances = []       # (month_iso, wallet_name, balance)
    budgets = []        # (income_name, month_iso, amount)
    paylater = []       # (item, monthly, first, last, note)
    loans = []          # (person, lender, installment, [(period, paid)...])

    # --- Budgets (Setup: name in col 9, Jan..Dec in cols 10..21) ---
    ws = wb["Setup"]
    r = 3
    while True:
        name = text(ws.cell(r, 9).value)
        if not name:
            break
        for m in range(12):
            amt = to_int(ws.cell(r, 10 + m).value)
            budgets.append((name, f"{YEAR}-{m + 1:02d}-01", amt or 0))
        r += 1

    # --- Wallet balances (Dashboard) ---
    ws = wb["Dashboard"]
    col_wallet = {}
    for c in range(1, ws.max_column + 1):
        nm = text(ws.cell(2, c).value)
        if nm:
            col_wallet[c] = nm
    seen_months: set[str] = set()
    for r in range(3, ws.max_row + 1):
        label = text(ws.cell(r, 1).value)
        if not label:
            continue
        if label.lower().startswith("tahun"):
            month_iso = f"{YEAR - 1}-12-01"
        elif label in MONTHS:
            month_iso = f"{YEAR}-{MONTHS.index(label) + 1:02d}-01"
        else:
            continue
        # The Dashboard sheet repeats the month list in a second block lower down
        # (different data, same columns). Stop once the first block is consumed.
        if month_iso in seen_months:
            break
        seen_months.add(month_iso)
        for c, wname in col_wallet.items():
            if wname.lower() == "wallet" or wname.lower() == "networth":
                continue
            bal = to_int(ws.cell(r, c).value)
            if bal is not None:
                balances.append((month_iso, wname, bal))

    # --- Monthly transactions ---
    for mi, mname in enumerate(MONTHS):
        if mname not in wb.sheetnames:
            continue
        ws = wb[mname]
        for r in range(5, ws.max_row + 1):
            # Expenses A-E
            d = to_date(ws.cell(r, 1).value)
            amt = to_int(ws.cell(r, 4).value)
            if d and amt:
                transactions.append(dict(
                    occurred_on=d, type="expense", amount=amt,
                    description=text(ws.cell(r, 3).value),
                    cat=("expense", text(ws.cell(r, 2).value)),
                    src=text(ws.cell(r, 5).value), dst=None))
            # Income G-K
            d = to_date(ws.cell(r, 7).value)
            amt = to_int(ws.cell(r, 10).value)
            if d and amt:
                transactions.append(dict(
                    occurred_on=d, type="income", amount=amt,
                    description=text(ws.cell(r, 9).value),
                    cat=("income", text(ws.cell(r, 8).value)),
                    src=None, dst=text(ws.cell(r, 11).value)))
            # Saving/Investment M-Q
            d = to_date(ws.cell(r, 13).value)
            amt = to_int(ws.cell(r, 15).value)
            to_bucket = text(ws.cell(r, 17).value)
            if d and amt and to_bucket:
                kind = "investment" if to_bucket in INVESTMENT_BUCKETS else "saving"
                transactions.append(dict(
                    occurred_on=d, type=kind, amount=amt,
                    description=text(ws.cell(r, 14).value),
                    cat=(kind, to_bucket),
                    src=text(ws.cell(r, 16).value), dst=None))
            # Transfer S-V
            d = to_date(ws.cell(r, 19).value)
            amt = to_int(ws.cell(r, 22).value)
            if d and amt:
                transactions.append(dict(
                    occurred_on=d, type="transfer", amount=amt,
                    description=None, cat=None,
                    src=text(ws.cell(r, 20).value), dst=text(ws.cell(r, 21).value)))

    # --- Paylater ---
    ws = wb["Cicilan Paylater"]
    r = 3
    while True:
        item = text(ws.cell(r, 1).value)
        if not item:
            break
        first = min(12, max(1, int(to_int(ws.cell(r, 3).value) or 1)))
        last = min(12, max(1, int(to_int(ws.cell(r, 4).value) or 1)))
        paylater.append((
            item,
            to_int(ws.cell(r, 2).value) or 0,
            date(2026, first, 1).isoformat(),
            date(2026, last, 1).isoformat(),
            None,
        ))
        r += 1

    # --- Loans (read from the spreadsheet's "Pinjol Orang" sheet) ---
    ws = wb["Pinjol Orang"]
    r = 4
    while True:
        person = text(ws.cell(r, 1).value)
        if not person:
            # allow a couple of blank rows then stop
            if not any(text(ws.cell(r + k, 1).value) for k in range(1, 3)):
                break
            r += 1
            continue
        lender = text(ws.cell(r, 2).value)
        installment = to_int(ws.cell(r, 3).value) or 0
        periods = []
        for p in range(24):  # columns are Jan 2026 .. Dec 2027
            val = truthy(ws.cell(r, 4 + p).value)
            if val is None:
                continue
            month = date(2026 + p // 12, p % 12 + 1, 1).isoformat()
            periods.append((month, val))
        loans.append((person, lender, installment, periods))
        r += 1

    return transactions, balances, budgets, paylater, loans


# ----------------------------- db load -----------------------------

def _split_sql(sql: str):
    """Split SQL into statements, respecting $$…$$ blocks (function bodies) and
    -- line comments. We only split on ; that is outside a dollar-quoted block."""
    stmts, buf = [], []
    i, n = 0, len(sql)
    in_dollar = False
    while i < n:
        two = sql[i:i + 2]
        if not in_dollar and two == "--":  # line comment (outside a function body)
            j = sql.find("\n", i)
            i = n if j == -1 else j
            continue
        if two == "$$":  # toggle dollar-quoted body
            in_dollar = not in_dollar
            buf.append("$$")
            i += 2
            continue
        ch = sql[i]
        if ch == ";" and not in_dollar:
            stmts.append("".join(buf))
            buf = []
            i += 1
            continue
        buf.append(ch)
        i += 1
    if "".join(buf).strip():
        stmts.append("".join(buf))
    return [s.strip() for s in stmts if s.strip()]


def run_sql_file(conn, path: Path):
    """Execute a .sql file statement-by-statement, ignoring 'already exists'."""
    for stmt in _split_sql(path.read_text(encoding="utf-8")):
        try:
            with conn.cursor() as cur:
                cur.execute(stmt)
        except Exception as e:  # noqa: BLE001
            if "already exists" in str(e).lower():
                continue
            raise


def connect_db():
    import os
    env = load_env(ROOT / ".env.local")
    db_url = os.environ.get("DATABASE_URL") or env.get("DATABASE_URL")
    if not db_url:
        sys.exit("DATABASE_URL not set (env or .env.local).")
    try:
        import psycopg
    except ImportError:
        sys.exit('psycopg missing. Run: python -m pip install "psycopg[binary]" openpyxl')
    return psycopg.connect(db_url, autocommit=True)


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # Windows consoles default to cp1252
    except Exception:
        pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--migrate-only", action="store_true",
                    help="apply schema + seed only (no truncate, no data import) — safe on a live DB")
    args = ap.parse_args()

    # Apply schema/seed without touching existing data.
    if args.migrate_only:
        conn = connect_db()
        print("Applying schema + seed (no data import)…")
        run_sql_file(conn, ROOT / "supabase" / "migrations" / "0001_init.sql")
        run_sql_file(conn, ROOT / "supabase" / "seed.sql")
        conn.close()
        print("✅ Migration applied (your data is untouched).")
        return

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        sys.exit(f"Spreadsheet not found: {xlsx}")

    txns, balances, budgets, paylater, loans = parse_workbook(xlsx)

    print("Parsed:")
    print(f"  transactions : {len(txns)}")
    by_type = {}
    for t in txns:
        by_type[t["type"]] = by_type.get(t["type"], 0) + 1
    for k, v in sorted(by_type.items()):
        print(f"      {k:11}: {v}")
    print(f"  balances     : {len(balances)}")
    print(f"  budgets      : {len(budgets)}")
    print(f"  paylater     : {len(paylater)}")
    print(f"  loans        : {len(loans)} ({sum(len(l[3]) for l in loans)} payment periods)")

    if args.dry_run:
        print("\n--dry-run: nothing written.")
        return

    conn = connect_db()
    print("\nConnected. Ensuring schema + seed…")
    run_sql_file(conn, ROOT / "supabase" / "migrations" / "0001_init.sql")
    run_sql_file(conn, ROOT / "supabase" / "seed.sql")

    cur = conn.cursor()

    # Reference lookups (create-on-the-fly for anything new in the logs)
    cur.execute("select id, name from wallets")
    wallet_id = {n: i for i, n in cur.fetchall()}
    cur.execute("select id, kind, name from categories")
    cat_id = {(k, n): i for i, k, n in cur.fetchall()}

    def get_wallet(name):
        if not name:
            return None
        if name not in wallet_id:
            cur.execute(
                "insert into wallets(name, sort_order) values (%s, "
                "(select coalesce(max(sort_order),0)+1 from wallets)) returning id", (name,))
            wallet_id[name] = cur.fetchone()[0]
        return wallet_id[name]

    def get_cat(kind, name):
        if not name:
            return None
        key = (kind, name)
        if key not in cat_id:
            cur.execute("insert into categories(kind, name) values (%s,%s) returning id", (kind, name))
            cat_id[key] = cur.fetchone()[0]
        return cat_id[key]

    print("Clearing historical tables…")
    cur.execute("truncate transactions, wallet_balances, budgets, paylater_items, "
                "paylater_payments, loan_payments, loans, monthly_wallet_delta restart identity")

    print("Inserting transactions…")
    rows = []
    for t in txns:
        cat = get_cat(*t["cat"]) if t["cat"] and t["cat"][1] else None
        rows.append((t["occurred_on"], t["type"], t["amount"], t["description"],
                     cat, get_wallet(t["src"]), get_wallet(t["dst"])))
    cur.executemany(
        "insert into transactions(occurred_on,type,amount,description,category_id,"
        "source_wallet_id,dest_wallet_id) values (%s,%s,%s,%s,%s,%s,%s)", rows)

    print("Inserting balances…")
    cur.executemany(
        "insert into wallet_balances(month, wallet_id, balance) values (%s,%s,%s) "
        "on conflict (month, wallet_id) do update set balance=excluded.balance",
        [(m, get_wallet(w), b) for m, w, b in balances if get_wallet(w)])

    print("Inserting budgets…")
    brows = []
    for name, month, amt in budgets:
        cid = get_cat("income", name)
        if cid:
            brows.append((cid, month, amt))
    cur.executemany(
        "insert into budgets(category_id, month, amount) values (%s,%s,%s) "
        "on conflict (category_id, month) do update set amount=excluded.amount", brows)

    print("Inserting paylater…")
    cur.executemany(
        "insert into paylater_items(item, monthly_amount, first_month_date, last_month_date, note) "
        "values (%s,%s,%s,%s,%s)", paylater)

    print("Inserting loans…")
    for person, lender, installment, periods in loans:
        cur.execute(
            "insert into loans(person, lender, installment) values (%s,%s,%s) returning id",
            (person, lender, installment))
        loan_id = cur.fetchone()[0]
        if periods:
            cur.executemany(
                "insert into loan_payments(loan_id, period_month, paid) values (%s,%s,%s)",
                [(loan_id, month, paid) for month, paid in periods])

    conn.commit()
    conn.close()
    print("\n✅ Import complete.")


if __name__ == "__main__":
    main()
